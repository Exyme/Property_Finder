# data_formatter.py
# Custom data filtering, sorting, and Excel export module
# 
# This module sits between distance_calculator and email_notifier
# Creates a formatted Excel file with custom filters and sorting

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config import CONFIG, get_type_aware_filename


def evaluate_condition(row, condition):
    """
    Evaluate a single filter condition against a row.
    
    Args:
        row: pandas Series (a single row from DataFrame)
        condition: dict with 'column', 'op', 'value' keys
        
    Returns:
        bool: True if condition is met, False otherwise
    """
    column = condition.get('column')
    op = condition.get('op')
    value = condition.get('value')
    
    if column not in row.index:
        print(f"⚠️  Warning: Column '{column}' not found in data")
        return True  # Skip condition if column doesn't exist
    
    cell_value = row[column]
    
    # Handle None/NaN values
    if pd.isna(cell_value):
        if op == 'is_empty':
            return value == True
        elif op == 'is_not_empty':
            return value == False
        else:
            return False  # NaN fails most comparisons
    
    # Comparison operators
    try:
        if op == '<=':
            return float(cell_value) <= float(value)
        elif op == '>=':
            return float(cell_value) >= float(value)
        elif op == '<':
            return float(cell_value) < float(value)
        elif op == '>':
            return float(cell_value) > float(value)
        elif op == '==':
            return cell_value == value
        elif op == '!=':
            return cell_value != value
        elif op == 'contains':
            return str(value).lower() in str(cell_value).lower()
        elif op == 'startswith':
            return str(cell_value).lower().startswith(str(value).lower())
        elif op == 'is_empty':
            return (pd.isna(cell_value) or str(cell_value).strip() == '') == value
        elif op == 'is_not_empty':
            return (not pd.isna(cell_value) and str(cell_value).strip() != '') == value
        else:
            print(f"⚠️  Warning: Unknown operator '{op}'")
            return True
    except (ValueError, TypeError) as e:
        # If comparison fails (e.g., comparing string to number), return False
        return False


def evaluate_filter(row, filter_item):
    """
    Evaluate a filter item which can be a condition or an OR group.
    
    Args:
        row: pandas Series
        filter_item: dict - either a condition or {'OR': [conditions]}
        
    Returns:
        bool: True if filter passes, False otherwise
    """
    if 'OR' in filter_item:
        # OR group - at least one condition must be true
        or_conditions = filter_item['OR']
        return any(evaluate_filter(row, cond) for cond in or_conditions)
    else:
        # Single condition
        return evaluate_condition(row, filter_item)


def apply_filters(df, filter_config):
    """
    Apply filters to a DataFrame.
    
    All top-level filters are combined with AND.
    Use {'OR': [...]} to group conditions with OR logic.
    
    Args:
        df: pandas DataFrame
        filter_config: list of filter conditions
        
    Returns:
        pandas DataFrame: Filtered data
    """
    if not filter_config:
        return df
    
    # Create a mask for rows that pass all filters
    mask = pd.Series([True] * len(df), index=df.index)
    
    for filter_item in filter_config:
        # Evaluate each filter for each row
        filter_mask = df.apply(lambda row: evaluate_filter(row, filter_item), axis=1)
        mask = mask & filter_mask
    
    filtered_df = df[mask].copy()
    return filtered_df


def apply_sorting(df, sort_config):
    """
    Apply multi-column sorting to a DataFrame.
    
    Args:
        df: pandas DataFrame
        sort_config: list of {'column': str, 'ascending': bool}
        
    Returns:
        pandas DataFrame: Sorted data
    """
    if not sort_config:
        return df
    
    # Extract column names and ascending flags
    columns = []
    ascending = []
    
    for sort_item in sort_config:
        col = sort_item.get('column')
        asc = sort_item.get('ascending', True)
        
        if col in df.columns:
            columns.append(col)
            ascending.append(asc)
        else:
            print(f"⚠️  Warning: Sort column '{col}' not found in data")
    
    if columns:
        df = df.sort_values(by=columns, ascending=ascending)
    
    return df


def detect_url_columns(df):
    """
    Detect columns that likely contain URLs based on content.
    
    Args:
        df: pandas DataFrame
        
    Returns:
        list: Column names that appear to contain URLs
    """
    url_columns = []
    for col in df.columns:
        series = df[col].dropna()
        if len(series) == 0:
            continue
        # Check if >50% of values look like URLs
        url_like = series.apply(
            lambda x: isinstance(x, str) and x.strip().startswith(('http://', 'https://', 'www.'))
        )
        if url_like.mean() > 0.5:
            url_columns.append(col)
    return url_columns


def export_to_excel(df, output_path, format_config):
    """
    Export DataFrame to Excel with formatting.
    
    Args:
        df: pandas DataFrame
        output_path: str - path to save the Excel file
        format_config: dict with formatting options
        
    Returns:
        str: Path to the created Excel file
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Properties"
    
    # Detect URL columns based on content
    url_columns = detect_url_columns(df)
    print(f"   🔗 Detected URL columns: {url_columns}")  # Debug output
    
    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            col_name = df.columns[c_idx - 1] if c_idx <= len(df.columns) else None
            
            # Check if this is a URL column with a URL value (for data rows only)
            is_url_cell = False
            url_value = None
            if r_idx > 1 and col_name in url_columns and value and isinstance(value, str):
                url_str = str(value).strip()
                if url_str.startswith(('http://', 'https://', 'www.')):
                    is_url_cell = True
                    url_value = url_str
                    if url_value.startswith('www.'):
                        url_value = 'https://' + url_value
            
            # Create cell
            cell = ws.cell(row=r_idx, column=c_idx)
            
            # Set value first
            cell.value = value
            
            # Then set hyperlink if needed
            if is_url_cell:
                cell.hyperlink = url_value
                cell.font = Font(color="0563C1", underline="single")
                cell.value = url_value            
            
            # Format header row (must come after hyperlink to avoid overwriting)
            if r_idx == 1:
                if format_config.get('bold_header', True):
                    # Only apply bold if not already a hyperlink
                    if not is_url_cell:
                        cell.font = Font(bold=True)
                    else:
                        # For header hyperlinks, combine bold with hyperlink style
                        cell.font = Font(bold=True, color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal='center')    
    
    # Freeze header row
    if format_config.get('freeze_header', True):
        ws.freeze_panes = 'A2'
    
    # Auto-filter
    if format_config.get('auto_filter', True) and len(df) > 0:
        ws.auto_filter.ref = ws.dimensions
    
    # Auto-adjust column widths
    if format_config.get('auto_column_width', True):
        for col_idx, column in enumerate(df.columns, 1):
            # Calculate max width based on header and first 100 rows
            max_length = len(str(column))
            for row in ws.iter_rows(min_row=2, max_row=min(102, len(df) + 1), 
                                     min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Set column width (with some padding, max 50)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Save the workbook
    try:
        wb.save(output_path)
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")
        return None
    
    return output_path


def format_and_export(args, input_csv_path=None):
    """
    Main entry point for the data formatter.
    
    Applies filters and sorting from config.py, then exports to Excel.
    
    Args:
        args: Argument object with output_dir, test_mode, file_suffix, property_type attributes
        input_csv_path: Path to the input CSV (defaults to type-aware property_listings_with_distances.csv)
        
    Returns:
        str or None: Path to the created Excel file, or None if disabled/failed
    """
    # Get configuration
    formatter_config = CONFIG.get('data_formatter', {})
    
    # Check if enabled
    if not formatter_config.get('enabled', True):
        print("📊 Data formatter is disabled in config.py")
        return None
    
    # Get settings from args
    output_dir = getattr(args, 'output_dir', 'output')
    test_mode = getattr(args, 'test_mode', False)
    file_suffix = getattr(args, 'file_suffix', '')
    property_type = getattr(args, 'property_type', 'rental')  # Get property_type from args
    
    # Auto-disable in test mode
    if test_mode:
        print("🧪 TEST MODE: Data formatter is disabled")
        return None
    
    # Determine input file (type-aware, with backward compatibility)
    if input_csv_path is None:
        # Try type-aware filename first
        input_filename = get_type_aware_filename('property_listings_with_distances', property_type, file_suffix)
        input_csv_path = os.path.join(output_dir, input_filename)
        # If not found, try old naming for backward compatibility
        if not os.path.exists(input_csv_path) and property_type == 'rental':
            old_input_csv_path = os.path.join(output_dir, f'property_listings_with_distances{file_suffix}.csv')
            if os.path.exists(old_input_csv_path):
                input_csv_path = old_input_csv_path
    
    if not os.path.exists(input_csv_path):
        print(f"❌ Error: Input file not found: {input_csv_path}")
        return None
    
    print("="*70)
    print("DATA FORMATTER - CUSTOM FILTERING & EXCEL EXPORT")
    print("="*70)
    print()
    
    # Load the data
    print(f"📂 Loading: {input_csv_path}")
    try:
        df = pd.read_csv(input_csv_path)
    except Exception as e:
        print(f"❌ Error loading CSV: {e}")
        return None
    initial_count = len(df)
    print(f"   Loaded {initial_count} properties")
    
    # Apply filters
    filter_config = formatter_config.get('filters', [])
    
    # For sales properties, remove price-related filters
    if property_type == 'sales' and filter_config:
        def remove_price_filters(filters):
            """Recursively remove price filters from filter config."""
            result = []
            for filter_item in filters:
                if 'OR' in filter_item:
                    # Handle OR groups - remove price filters from within OR group
                    or_conditions = filter_item['OR']
                    filtered_or = [cond for cond in or_conditions if cond.get('column') != 'price']
                    if filtered_or:  # Only keep OR group if it has remaining conditions
                        result.append({'OR': filtered_or})
                else:
                    # Single condition - skip if it's a price filter
                    if filter_item.get('column') != 'price':
                        result.append(filter_item)
            return result
        
        filter_config = remove_price_filters(filter_config)
        print(f"   [SALES] Removed price filters from data formatter")
    
    if filter_config:
        print(f"\n🔍 Applying {len(filter_config)} filter(s)...")
        df = apply_filters(df, filter_config)
        filtered_count = len(df)
        removed_count = initial_count - filtered_count
        print(f"   Filtered: {filtered_count} properties ({removed_count} removed)")
    
    # Apply sorting
    sort_config = formatter_config.get('sort_by', [])
    
    # For sales properties, remove price from sort_by
    if property_type == 'sales' and sort_config:
        sort_config = [s for s in sort_config if s.get('column') != 'price']
        if sort_config:
            print(f"   [SALES] Removed price from sort_by")
    if sort_config:
        sort_cols = [s.get('column') for s in sort_config]
        print(f"\n📊 Sorting by: {', '.join(sort_cols)}")
        df = apply_sorting(df, sort_config)
    
    # Check if any properties remain
    if len(df) == 0:
        print("\n⚠️  No properties match the filter criteria")
        return None
    
    # Export to Excel (type-aware filename)
    base_output_filename = formatter_config.get('output_filename', 'property_listings_filtered.xlsx')
    # Extract base name without extension
    base_name = base_output_filename.replace('.xlsx', '').replace('.xls', '')
    # Use type-aware filename generation
    output_filename = get_type_aware_filename(base_name, property_type, file_suffix, extension='xlsx')
    output_path = os.path.join(output_dir, output_filename)
    
    print(f"\n📁 Exporting to Excel...")
    excel_path = export_to_excel(df, output_path, formatter_config)
    if excel_path:
        print(f"   ✅ Saved: {excel_path}")
        print(f"   Properties: {len(df)}")
    else:
        print("   ❌ Export failed")
        return None
    
    # Summary
    print()
    print("="*70)
    print("✅ DATA FORMATTER COMPLETE")
    print("="*70)
    print(f"   Input: {initial_count} properties")
    print(f"   Output: {len(df)} properties (filtered)")
    print(f"   File: {output_filename}")
    print()
    
    return excel_path


# For standalone testing (remove or comment out if integrating elsewhere)
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Run data formatter standalone")
    parser.add_argument('--output_dir', default='output', help='Output directory')
    parser.add_argument('--test_mode', action='store_true', help='Enable test mode')
    parser.add_argument('--file_suffix', default='', help='File suffix')
    parser.add_argument('--input_csv', help='Path to input CSV')
    args = parser.parse_args()
    format_and_export(args, input_csv_path=args.input_csv)